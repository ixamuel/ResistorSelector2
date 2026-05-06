"""
extract_from_xlsm.py
====================
Extract resistor data from a Panasonic XLSM (macro-enabled workbook) file,
save as clean XLSX, and optionally run the existing pipeline to regenerate
resistors_compact.json and data.js.

Dual Strategy:
  - Strategy A (Preferred): Use openpyxl to read the XLSM directly.
    Sheet protection only blocks UI editing, not programmatic access.
  - Strategy B (Fallback): Extract XLSM as ZIP archive, parse worksheet XML
    directly, reading shared strings. This bypasses ALL protection mechanisms.

Usage:
    python extract_from_xlsm.py
    python extract_from_xlsm.py --skip-pipeline
    python extract_from_xlsm.py --force-zip
    python extract_from_xlsm.py path/to/file.xlsm --output-xlsx Resistor_DB.xlsx
"""

import sys
# Force UTF-8 output for Windows terminals
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import os
import sys
import json
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

# Try openpyxl first (Strategy A)
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Try pandas (needed for the existing pipeline)
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


# ── Namespace for Office Open XML ──────────────────────────────────────────
NS = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
NS_REL = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}


# ═══════════════════════════════════════════════════════════════════════════
#  STRATEGY A  —  openpyxl direct read
# ═══════════════════════════════════════════════════════════════════════════

def extract_sheet_via_openpyxl(xlsm_path):
    """
    Use openpyxl to read the XLSM directly.
    Sheet protection only blocks UI editing, so this usually works.
    Returns a list of lists (rows), where the first row is the header.
    """
    print("  [Strategy A] Reading XLSM with openpyxl (data_only=True)...")
    wb = openpyxl.load_workbook(xlsm_path, data_only=True, keep_vba=False)

    # Try to find the most data-rich sheet
    sheet_names = wb.sheetnames
    print(f"  Sheets found: {sheet_names}")

    # Prefer the sheet with the most rows (the data sheet)
    best_sheet = None
    best_rows = 0
    for name in sheet_names:
        ws = wb[name]
        row_count = ws.max_row or 0
        if row_count > best_rows:
            best_rows = row_count
            best_sheet = name

    if best_sheet is None:
        wb.close()
        raise ValueError("No sheets found in workbook")

    print(f"  Using sheet: '{best_sheet}' ({best_rows} rows)")
    ws = wb[best_sheet]

    # Convert to list of lists
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        data.append(list(row))

    wb.close()
    return data


# ═══════════════════════════════════════════════════════════════════════════
#  STRATEGY B  —  ZIP extraction + XML parsing
# ═══════════════════════════════════════════════════════════════════════════

def _parse_shared_strings(zip_file):
    """Parse sharedStrings.xml from the XLSM archive."""
    if 'xl/sharedStrings.xml' not in zip_file.namelist():
        return []

    ss_xml = zip_file.read('xl/sharedStrings.xml')
    root = ET.fromstring(ss_xml)

    strings = []
    for si in root.findall('.//s:si', NS):
        # Collect all <t> elements (text runs can be split across multiple <t>)
        text_parts = []
        for t in si.iter(f'{{{NS["s"]}}}t'):
            if t.text:
                text_parts.append(t.text)
        strings.append(''.join(text_parts))

    return strings


def _find_data_sheet_xml(zip_file):
    """
    Find the worksheet XML file that contains the resistor data.
    Looks at xl/workbook.xml to find sheet names, then checks relationships
    to map sheet names to file paths.
    """
    # Read workbook.xml to get sheet names and their relationship IDs
    wb_xml = zip_file.read('xl/workbook.xml')
    wb_root = ET.fromstring(wb_xml)

    sheets = []
    for sheet_elem in wb_root.findall('.//s:sheet', NS):
        name = sheet_elem.get('name')
        r_id = sheet_elem.get(f'{{{NS_REL["r"]}}}id')
        sheets.append((name, r_id))

    if not sheets:
        raise ValueError("No sheets found in workbook.xml")

    # Read relationships to map r:id → target file path
    rels_xml = zip_file.read('xl/_rels/workbook.xml.rels')
    rels_root = ET.fromstring(rels_xml)
    rels_ns = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
    rel_map = {}
    for rel_elem in rels_root:
        r_id = rel_elem.get('Id')
        target = rel_elem.get('Target')
        if r_id and target:
            rel_map[r_id] = target

    # Build sheet name → file path mapping
    sheet_files = []
    for name, r_id in sheets:
        target = rel_map.get(r_id, '')
        # Target is relative to xl/ directory
        if target:
            path = f'xl/{target}' if not target.startswith('/') else target[1:]
            sheet_files.append((name, path))

    print(f"  Sheets found: {[s[0] for s in sheet_files]}")

    # Prefer the sheet with the most rows (check file sizes as heuristic)
    best_sheet = None
    best_size = 0
    for name, path in sheet_files:
        if path in zip_file.namelist():
            info = zip_file.getinfo(path)
            if info.file_size > best_size:
                best_size = info.file_size
                best_sheet = (name, path)

    if best_sheet is None:
        # Fallback: just use first worksheet
        ws_files = sorted([f for f in zip_file.namelist()
                          if f.startswith('xl/worksheets/') and f.endswith('.xml')])
        if not ws_files:
            raise ValueError("No worksheet XML files found in XLSM archive")
        best_sheet = (Path(ws_files[0]).stem, ws_files[0])

    print(f"  Using sheet: '{best_sheet[0]}'")
    return best_sheet[1]


def extract_sheet_via_zip(xlsm_path):
    """
    Extract XLSM as ZIP and parse the worksheet XML directly.
    This bypasses all protection mechanisms (sheet passwords, VBA locks, etc.).
    Returns a list of lists (rows), where the first row is the header.
    """
    print("  [Strategy B] Extracting XLSM as ZIP and parsing XML...")

    with zipfile.ZipFile(xlsm_path, 'r') as z:
        # List all files for debugging
        all_files = z.namelist()
        print(f"  Archive contains {len(all_files)} files")

        # Parse shared strings
        shared_strings = _parse_shared_strings(z)
        print(f"  Shared strings loaded: {len(shared_strings)}")

        # Find the data sheet
        sheet_path = _find_data_sheet_xml(z)
        sheet_xml = z.read(sheet_path)

    # Parse the sheet XML
    root = ET.fromstring(sheet_xml)

    # Collect all rows with their row numbers
    rows = {}
    for row_elem in root.findall('.//s:row', NS):
        row_num = int(row_elem.get('r', 0))
        cells = {}

        for cell in row_elem.findall('s:c', NS):
            cell_ref = cell.get('r', '')
            cell_type = cell.get('t', '')
            cell_value_elem = cell.find('s:v', NS)
            raw_value = cell_value_elem.text if cell_value_elem is not None else ''

            # Resolve value
            if cell_type == 's' and raw_value:
                # Shared string reference
                idx = int(raw_value)
                value = shared_strings[idx] if idx < len(shared_strings) else raw_value
            elif cell_type == 'b':
                value = raw_value  # boolean
            elif cell_type == 'e':
                value = raw_value  # error
            else:
                value = raw_value  # number or inline string

            cells[cell_ref] = value

        if cells:
            rows[row_num] = cells

    if not rows:
        raise ValueError("No data rows found in worksheet XML")

    # Convert to column-based ordering
    # Collect all column references in order
    all_col_refs = []
    for row_num in sorted(rows.keys()):
        for cell_ref in rows[row_num]:
            # Extract column letters (e.g., 'A', 'AB', 'C')
            col_letters = ''.join(c for c in cell_ref if c.isalpha())
            if col_letters and col_letters not in all_col_refs:
                all_col_refs.append(col_letters)

    # Build the data matrix
    data = []
    for row_num in sorted(rows.keys()):
        row_data = []
        for col_ref in all_col_refs:
            # Find the cell with this column reference in this row
            cell_key = f"{col_ref}{row_num}"
            row_data.append(rows[row_num].get(cell_key, ''))
        data.append(row_data)

    return data


# ═══════════════════════════════════════════════════════════════════════════
#  SAVE AS XLSX
# ═══════════════════════════════════════════════════════════════════════════

def save_as_xlsx(data, output_path):
    """Save extracted data as a clean XLSX file."""
    if not data:
        print("  Warning: No data to save!")
        return

    print(f"  Saving clean XLSX to: {output_path}")

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resistor Data"
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Try to convert numeric strings to numbers
                if isinstance(value, str):
                    # Try int first
                    try:
                        cell.value = int(value)
                        continue
                    except (ValueError, TypeError):
                        pass
                    # Try float
                    try:
                        cell.value = float(value)
                        continue
                    except (ValueError, TypeError):
                        pass
                cell.value = value
        wb.save(output_path)
        print(f"  Saved {len(data)} rows using openpyxl")
    elif HAS_PANDAS:
        df = pd.DataFrame(data[1:], columns=data[0] if data else None)
        df.to_excel(output_path, index=False)
        print(f"  Saved {len(data)} rows using pandas")
    else:
        print("  ERROR: Neither openpyxl nor pandas available. Cannot save XLSX.")
        print("  Install with: pip install openpyxl pandas")
        return False

    return True


# ═══════════════════════════════════════════════════════════════════════════
#  PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def run_pipeline(xlsx_path, json_output='resistors_compact.json', data_js='data.js'):
    """Run the existing pipeline: XLSX → JSON → data.js"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, script_dir)

    # Step 1: XLSX → JSON (compact_resistors_v2.create_compact_json)
    print("\n── Step 1: Converting XLSX to compact JSON ──")
    try:
        from compact_resistors_v2 import create_compact_json
        create_compact_json(xlsx_path, json_output)
    except Exception as e:
        print(f"  ERROR in compact_resistors_v2: {e}")
        import traceback
        traceback.print_exc()
        return False

    # Step 2: JSON → data.js (update_data_app.update_data_js)
    print("\n── Step 2: Generating data.js ──")
    try:
        from update_data_app import update_data_js
        update_data_js(json_output, data_js)
    except Exception as e:
        print(f"  ERROR in update_data_app: {e}")
        import traceback
        traceback.print_exc()
        return False

    print("\n[DONE] Pipeline complete!")
    print(f"   {xlsx_path}")
    print(f"   → {json_output}")
    print(f"   → {data_js}")
    return True


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Extract resistor data from Panasonic XLSM and update the database'
    )
    parser.add_argument(
        'xlsm_file',
        nargs='?',
        default='PanasonicIndustry_Chip_Resistor_SelectionTool.xlsm',
        help='Path to the XLSM file (default: PanasonicIndustry_Chip_Resistor_SelectionTool.xlsm)'
    )
    parser.add_argument(
        '--output-xlsx',
        default='Resistor DB.xlsx',
        help='Output XLSX file path (default: Resistor DB.xlsx)'
    )
    parser.add_argument(
        '--json-output',
        default='resistors_compact.json',
        help='Output JSON file path (default: resistors_compact.json)'
    )
    parser.add_argument(
        '--data-js',
        default='data.js',
        help='Output data.js file path (default: data.js)'
    )
    parser.add_argument(
        '--skip-pipeline',
        action='store_true',
        help='Skip the JSON/data.js generation pipeline (just extract to XLSX)'
    )
    parser.add_argument(
        '--force-zip',
        action='store_true',
        help='Force ZIP extraction method instead of openpyxl'
    )

    args = parser.parse_args()

    # ── Validate input ──────────────────────────────────────────────────
    if not os.path.exists(args.xlsm_file):
        print(f"[ERROR] XLSM file not found: {args.xlsm_file}")
        print(f"   Current directory: {os.getcwd()}")
        sys.exit(1)

    print(f"[INPUT]  {args.xlsm_file}")
    print(f"[OUTPUT] {args.output_xlsx}")
    print(f"   JSON:   {args.json_output}")
    print(f"   data.js: {args.data_js}")
    print()

    # ── Extract data ────────────────────────────────────────────────────
    data = None

    if args.force_zip:
        print("Forcing ZIP extraction method (--force-zip)...")
        data = extract_sheet_via_zip(args.xlsm_file)
    elif HAS_OPENPYXL:
        print("Attempting Strategy A (openpyxl)...")
        try:
            data = extract_sheet_via_openpyxl(args.xlsm_file)
        except Exception as e:
            print(f"  [WARN] Strategy A failed: {e}")
            print("  Falling back to Strategy B (ZIP extraction)...")
            data = extract_sheet_via_zip(args.xlsm_file)
    else:
        print("openpyxl not available. Using Strategy B (ZIP extraction)...")
        data = extract_sheet_via_zip(args.xlsm_file)

    if not data:
        print("[ERROR] No data extracted from XLSM")
        sys.exit(1)

    print(f"\n[OK] Extracted {len(data)} rows (including header)")
    if len(data) > 0:
        print(f"   Columns: {len(data[0])}")
        print(f"   Header:  {data[0][:6]}...")  # Show first 6 columns

    # ── Save as clean XLSX ──────────────────────────────────────────────
    print()
    success = save_as_xlsx(data, args.output_xlsx)
    if not success:
        sys.exit(1)

    # ── Run pipeline ────────────────────────────────────────────────────
    if not args.skip_pipeline:
        run_pipeline(args.output_xlsx, args.json_output, args.data_js)
    else:
        print(f"\n[DONE] Clean XLSX saved to: {args.output_xlsx}")
        print("   Run 'python compact_resistors_v2.py' to regenerate JSON/data.js.")

    # Show file sizes
    print()
    for fpath in [args.xlsm_file, args.output_xlsx, args.json_output, args.data_js]:
        if os.path.exists(fpath):
            size_kb = os.path.getsize(fpath) / 1024
            print(f"   {fpath}: {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
